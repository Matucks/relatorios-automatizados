import pandas as pd
import glob
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Configurações do servidor SMTP
SMTP_SERVER = 'smtp.seuservidor.com'
SMTP_PORT = 587
EMAIL_USER = 'seu.email@seuservidor.com'
EMAIL_PASS = 'sua_senha'
EMAIL_SUBJECT = 'Relatório Automático'
EMAIL_BODY = (
    "Bom dia,\n\n"
    "Segue em anexo o arquivo gerado automaticamente.\n\n"
    "Atenciosamente,\n"
    "Equipe de Automação"
)

# Mapeamento de categorias
categorias = {
    "BLAZER EV": "SUV elétrico",
    "BOLT": "SUV elétrico",
    "EQUINOX EV": "SUV elétrico",
    "MONTANA": "picape",
    "ONIX": "hatchback",
    "ONIX PLUS": "SEDAN",
    "S10": "picape",
    "SILVERADO": "picape grande",
    "SPIN": "minivan",
    "TRACKER": "SUV compacto"
}

# Definindo caminhos das pastas de entrada e saída
pasta_entrada = r"C:\Projeto\localizador\entrada"
caminho_saida = r"C:\Projeto\localizador\saida\LOCALIZADOR_GM.xlsx"

# Encontrando o arquivo Excel mais recente na pasta de entrada
arquivos_entrada = glob.glob(os.path.join(pasta_entrada, "*.xlsx"))
if not arquivos_entrada:
    raise FileNotFoundError("Nenhum arquivo Excel encontrado na pasta de entrada.")

# Selecionando o arquivo mais recente
arquivo_entrada = max(arquivos_entrada, key=os.path.getctime)

# Carregando a planilha de entrada
df = pd.read_excel(arquivo_entrada)

# Verificando se as colunas necessárias existem no dataframe
colunas_necessarias = ['Status veículo', 'Cor externa', 'Linha de produto', 'Opcionais', 'Código Modelo']
for coluna in colunas_necessarias:
    if coluna not in df.columns:
        raise ValueError(f"A coluna '{coluna}' não foi encontrada no arquivo de entrada.")

# Convertendo os nomes das colunas para strings
df.columns = df.columns.map(str)

# Criando o arquivo Excel de saída com openpyxl para aplicar formatação
with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
    produtos = df['Linha de produto'].unique()
    data_geracao = datetime.now().strftime("%d/%m/%Y")  # Data em formato PT-BR

    # Criando a aba índice
    workbook = writer.book
    worksheet_indice = workbook.create_sheet("Índice")
    
    # Configuração do índice
    worksheet_indice.merge_cells("A1:F1")
    cell_title = worksheet_indice["A1"]
    cell_title.value = "Menu de Modelos"
    cell_title.font = Font(size=16, bold=True, color="FFFFFF")
    cell_title.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    worksheet_indice.row_dimensions[1].height = 30

    # Adicionando cabeçalho para o índice
    header_row = 3
    headers = ["Modelo", "Descrição", "Data", "Status", "Categoria"]
    for col_num, header in enumerate(headers, start=1):
        cell = worksheet_indice.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(size=12, bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    worksheet_indice.row_dimensions[header_row].height = 20

    # Populando o índice com hiperlinks
    for i, produto in enumerate(produtos, start=header_row + 1):
        nome_aba = str(produto).replace(" ", "_")
        categoria = categorias.get(produto.upper(), "Categoria desconhecida")
        
        worksheet_indice.cell(row=i, column=1, value=produto).hyperlink = f"#{nome_aba}!A1"
        worksheet_indice.cell(row=i, column=1).font = Font(size=11, bold=True, color="0070C0")
        worksheet_indice.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center")
        worksheet_indice.cell(row=i, column=2, value="Descrição do modelo")
        worksheet_indice.cell(row=i, column=3, value=data_geracao)
        worksheet_indice.cell(row=i, column=4, value="Ativo")
        worksheet_indice.cell(row=i, column=5, value=categoria)
        
        # Aplicando cor de fundo alternada para linhas
        fill_color = "F2F2F2" if i % 2 == 0 else "FFFFFF"
        for col in range(1, 6):
            worksheet_indice.cell(row=i, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            worksheet_indice.cell(row=i, column=col).border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

    # Ajustando larguras das colunas
    column_widths = [40, 30, 15, 15, 20]
    for col_num, width in enumerate(column_widths, start=1):
        worksheet_indice.column_dimensions[get_column_letter(col_num)].width = width

    # Adicionando rodapé
    footer_row = len(produtos) + header_row + 2
    worksheet_indice.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=5)
    footer_cell = worksheet_indice.cell(row=footer_row, column=1, value="Este índice foi gerado automaticamente. Clique no modelo para acessar.")
    footer_cell.font = Font(size=10, italic=True)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    footer_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    worksheet_indice.row_dimensions[footer_row].height = 20

    # Gerando abas para cada produto
    for idx, produto in enumerate(produtos):
        df_produto = df[df['Linha de produto'] == produto]
        tabela_dinamica = pd.pivot_table(
            df_produto,
            index=['Linha de produto', 'Opcionais'],
            columns=['Cor externa', 'Status veículo'],
            values='Código Modelo',
            aggfunc='count',
            fill_value=0
        ).reset_index()
        
        tabela_dinamica.columns = tabela_dinamica.columns.map(str)
        nome_aba = str(produto).replace(' ', '_')
        tabela_dinamica.to_excel(writer, sheet_name=nome_aba, index=False)
        worksheet = writer.sheets[nome_aba]
        num_rows, num_cols = tabela_dinamica.shape
        tabela_ref = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
        nome_tabela = f"Table_{nome_aba}_{idx}"
        tabela = Table(displayName=nome_tabela, ref=tabela_ref)
        estilo_tabela = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tabela.tableStyleInfo = estilo_tabela
        worksheet.add_table(tabela)

        # Ajustar largura das colunas
        for col_idx, column_cells in enumerate(worksheet.columns, start=1):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

        # Adicionar botão de retorno ao índice
        botao_row = num_rows + 3
        cell_botao = worksheet.cell(row=botao_row, column=1, value="Voltar para o Índice")
        cell_botao.hyperlink = "#Índice!A1"
        cell_botao.font = Font(color="FFFFFF", bold=True)
        cell_botao.alignment = Alignment(horizontal="center", vertical="center")
        cell_botao.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        worksheet.row_dimensions[botao_row].height = 20

print(f"As tabelas dinâmicas e o índice foram gerados e salvos no arquivo de saída: {caminho_saida}")

# Função para envio de e-mail
def enviar_email_geral(lista_emails, caminho_anexo):
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
            servidor.starttls()
            servidor.login(EMAIL_USER, EMAIL_PASS)
            mensagem = MIMEMultipart()
            mensagem['From'] = EMAIL_USER
            mensagem['To'] = ', '.join(lista_emails)
            mensagem['Subject'] = EMAIL_SUBJECT
            mensagem.attach(MIMEText(EMAIL_BODY, 'plain'))
            with open(caminho_anexo, 'rb') as anexo:
                parte = MIMEBase('application', 'octet-stream')
                parte.set_payload(anexo.read())
                encoders.encode_base64(parte)
                parte.add_header(
                    'Content-Disposition',
                    f'attachment; filename={os.path.basename(caminho_anexo)}'
                )
                mensagem.attach(parte)
            servidor.sendmail(EMAIL_USER, lista_emails, mensagem.as_string())
            print("E-mail enviado com sucesso para os destinatários.")
    except Exception as e:
        print(f"Erro ao enviar o e-mail: {e}")

# Lista de destinatários
# Substituir pela lista de e-mails relevante quando necessário.
destinatarios = []

# Enviando o e-mail
enviar_email_geral(destinatarios, caminho_saida)

print("Processo finalizado com sucesso. Arquivo gerado e e-mails enviados.")
